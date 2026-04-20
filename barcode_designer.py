"""
barcode_designer.py – Design ONT nanopore barcode adapters for each fragment.

For each Part/Fragment, generates SnapGene .dna files for:
  - NBXX_TopStrand.dna         (40 nt, linear, Unmodified ends)
  - NBXX_BottomStrand.dna      (50 nt, linear, Unmodified ends)
  - XXXX_NBXX.dna              (annealed adapter with sticky ends, FivePrimePhosphorylated)
  - Ligation.dna               (full ligation product with native adapters)

Barcode assignment rules:
  - Each fragment's oh5 and oh3 overhangs each get one barcode (NB01–NB96)
  - If an overhang already has a barcode (from history_df), reuse it
  - No two fragments can share the same (oh5_nb, oh3_nb) combination
  - New barcodes assigned from the least-used NB number for even distribution
"""

from __future__ import annotations

import io
import zipfile
from collections import Counter
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd

from fragment_splitter import Fragment, Part
from dna_writer import build_dna_bytes, make_feature, make_multiseg_feature

# ---------------------------------------------------------------------------
# All 96 ONT barcode sequences
# ---------------------------------------------------------------------------

ALL_BARCODES: Dict[str, Dict[str, str]] = {
    "NB01": {"fwd": "CACAAAGACACCGACAACTTTCTT", "rev": "AAGAAAGTTGTCGGTGTCTTTGTG"},
    "NB02": {"fwd": "ACAGACGACTACAAACGGAATCGA", "rev": "TCGATTCCGTTTGTAGTCGTCTGT"},
    "NB03": {"fwd": "CCTGGTAACTGGGACACAAGACTC", "rev": "GAGTCTTGTGTCCCAGTTACCAGG"},
    "NB04": {"fwd": "TAGGGAAACACGATAGAATCCGAA", "rev": "TTCGGATTCTATCGTGTTTCCCTA"},
    "NB05": {"fwd": "AAGGTTACACAAACCCTGGACAAG", "rev": "CTTGTCCAGGGTTTGTGTAACCTT"},
    "NB06": {"fwd": "GACTACTTTCTGCCTTTGCGAGAA", "rev": "TTCTCGCAAAGGCAGAAAGTAGTC"},
    "NB07": {"fwd": "AAGGATTCATTCCCACGGTAACAC", "rev": "GTGTTACCGTGGGAATGAATCCTT"},
    "NB08": {"fwd": "ACGTAACTTGGTTTGTTCCCTGAA", "rev": "TTCAGGGAACAAACCAAGTTACGT"},
    "NB09": {"fwd": "AACCAAGACTCGCTGTGCCTAGTT", "rev": "AACTAGGCACAGCGAGTCTTGGTT"},
    "NB10": {"fwd": "GAGAGGACAAAGGTTTCAACGCTT", "rev": "AAGCGTTGAAACCTTTGTCCTCTC"},
    "NB11": {"fwd": "TCCATTCCCTCCGATAGATGAAAC", "rev": "GTTTCATCTATCGGAGGGAATGGA"},
    "NB12": {"fwd": "TCCGATTCTGCTTCTTTCTACCTG", "rev": "CAGGTAGAAAGAAGCAGAATCGGA"},
    "NB13": {"fwd": "AGAACGACTTCCATACTCGTGTGA", "rev": "TCACACGAGTATGGAAGTCGTTCT"},
    "NB14": {"fwd": "AACGAGTCTCTTGGGACCCATAGA", "rev": "TCTATGGGTCCCAAGAGACTCGTT"},
    "NB15": {"fwd": "AGGTCTACCTCGCTAACACCACTG", "rev": "CAGTGGTGTTAGCGAGGTAGACCT"},
    "NB16": {"fwd": "CGTCAACTGACAGTGGTTCGTACT", "rev": "AGTACGAACCACTGTCAGTTGACG"},
    "NB17": {"fwd": "ACCCTCCAGGAAAGTACCTCTGAT", "rev": "ATCAGAGGTACTTTCCTGGAGGGT"},
    "NB18": {"fwd": "CCAAACCCAACAACCTAGATAGGC", "rev": "GCCTATCTAGGTTGTTGGGTTTGG"},
    "NB19": {"fwd": "GTTCCTCGTGCAGTGTCAAGAGAT", "rev": "ATCTCTTGACACTGCACGAGGAAC"},
    "NB20": {"fwd": "TTGCGTCCTGTTACGAGAACTCAT", "rev": "ATGAGTTCTCGTAACAGGACGCAA"},
    "NB21": {"fwd": "GAGCCTCTCATTGTCCGTTCTCTA", "rev": "TAGAGAACGGACAATGAGAGGCTC"},
    "NB22": {"fwd": "ACCACTGCCATGTATCAAAGTACG", "rev": "CGTACTTTGATACATGGCAGTGGT"},
    "NB23": {"fwd": "CTTACTACCCAGTGAACCTCCTCG", "rev": "CGAGGAGGTTCACTGGGTAGTAAG"},
    "NB24": {"fwd": "GCATAGTTCTGCATGATGGGTTAG", "rev": "CTAACCCATCATGCAGAACTATGC"},
    "NB25": {"fwd": "GTAAGTTGGGTATGCAACGCAATG", "rev": "CATTGCGTTGCATACCCAACTTAC"},
    "NB26": {"fwd": "CATACAGCGACTACGCATTCTCAT", "rev": "ATGAGAATGCGTAGTCGCTGTATG"},
    "NB27": {"fwd": "CGACGGTTAGATTCACCTCTTACA", "rev": "TGTAAGAGGTGAATCTAACCGTCG"},
    "NB28": {"fwd": "TGAAACCTAAGAAGGCACCGTATC", "rev": "GATACGGTGCCTTCTTAGGTTTCA"},
    "NB29": {"fwd": "CTAGACACCTTGGGTTGACAGACC", "rev": "GGTCTGTCAACCCAAGGTGTCTAG"},
    "NB30": {"fwd": "TCAGTGAGGATCTACTTCGACCCA", "rev": "TGGGTCGAAGTAGATCCTCACTGA"},
    "NB31": {"fwd": "TGCGTACAGCAATCAGTTACATTG", "rev": "CAATGTAACTGATTGCTGTACGCA"},
    "NB32": {"fwd": "CCAGTAGAAGTCCGACAACGTCAT", "rev": "ATGACGTTGTCGGACTTCTACTGG"},
    "NB33": {"fwd": "CAGACTTGGTACGGTTGGGTAACT", "rev": "AGTTACCCAACCGTACCAAGTCTG"},
    "NB34": {"fwd": "GGACGAAGAACTCAAGTCAAAGGC", "rev": "GCCTTTGACTTGAGTTCTTCGTCC"},
    "NB35": {"fwd": "CTACTTACGAAGCTGAGGGACTGC", "rev": "GCAGTCCCTCAGCTTCGTAAGTAG"},
    "NB36": {"fwd": "ATGTCCCAGTTAGAGGAGGAAACA", "rev": "TGTTTCCTCCTCTAACTGGGACAT"},
    "NB37": {"fwd": "GCTTGCGATTGATGCTTAGTATCA", "rev": "TGATACTAAGCATCAATCGCAAGC"},
    "NB38": {"fwd": "ACCACAGGAGGACGATACAGAGAA", "rev": "TTCTCTGTATCGTCCTCCTGTGGT"},
    "NB39": {"fwd": "CCACAGTGTCAACTAGAGCCTCTC", "rev": "GAGAGGCTCTAGTTGACACTGTGG"},
    "NB40": {"fwd": "TAGTTTGGATGACCAAGGATAGCC", "rev": "GGCTATCCTTGGTCATCCAAACTA"},
    "NB41": {"fwd": "GGAGTTCGTCCAGAGAAGTACACG", "rev": "CGTGTACTTCTCTGGACGAACTCC"},
    "NB42": {"fwd": "CTACGTGTAAGGCATACCTGCCAG", "rev": "CTGGCAGGTATGCCTTACACGTAG"},
    "NB43": {"fwd": "CTTTCGTTGTTGACTCGACGGTAG", "rev": "CTACCGTCGAGTCAACAACGAAAG"},
    "NB44": {"fwd": "AGTAGAAAGGGTTCCTTCCCACTC", "rev": "GAGTGGGAAGGAACCCTTTCTACT"},
    "NB45": {"fwd": "GATCCAACAGAGATGCCTTCAGTG", "rev": "CACTGAAGGCATCTCTGTTGGATC"},
    "NB46": {"fwd": "GCTGTGTTCCACTTCATTCTCCTG", "rev": "CAGGAGAATGAAGTGGAACACAGC"},
    "NB47": {"fwd": "GTGCAACTTTCCCACAGGTAGTTC", "rev": "GAACTACCTGTGGGAAAGTTGCAC"},
    "NB48": {"fwd": "CATCTGGAACGTGGTACACCTGTA", "rev": "TACAGGTGTACCACGTTCCAGATG"},
    "NB49": {"fwd": "ACTGGTGCAGCTTTGAACATCTAG", "rev": "CTAGATGTTCAAAGCTGCACCAGT"},
    "NB50": {"fwd": "ATGGACTTTGGTAACTTCCTGCGT", "rev": "ACGCAGGAAGTTACCAAAGTCCAT"},
    "NB51": {"fwd": "GTTGAATGAGCCTACTGGGTCCTC", "rev": "GAGGACCCAGTAGGCTCATTCAAC"},
    "NB52": {"fwd": "TGAGAGACAAGATTGTTCGTGGAC", "rev": "GTCCACGAACAATCTTGTCTCTCA"},
    "NB53": {"fwd": "AGATTCAGACCGTCTCATGCAAAG", "rev": "CTTTGCATGAGACGGTCTGAATCT"},
    "NB54": {"fwd": "CAAGAGCTTTGACTAAGGAGCATG", "rev": "CATGCTCCTTAGTCAAAGCTCTTG"},
    "NB55": {"fwd": "TGGAAGATGAGACCCTGATCTACG", "rev": "CGTAGATCAGGGTCTCATCTTCCA"},
    "NB56": {"fwd": "TCACTACTCAACAGGTGGCATGAA", "rev": "TTCATGCCACCTGTTGAGTAGTGA"},
    "NB57": {"fwd": "GCTAGGTCAATCTCCTTCGGAAGT", "rev": "ACTTCCGAAGGAGATTGACCTAGC"},
    "NB58": {"fwd": "CAGGTTACTCCTCCGTGAGTCTGA", "rev": "TCAGACTCACGGAGGAGTAACCTG"},
    "NB59": {"fwd": "TCAATCAAGAAGGGAAAGCAAGGT", "rev": "ACCTTGCTTTCCCTTCTTGATTGA"},
    "NB60": {"fwd": "CATGTTCAACCAAGGCTTCTATGG", "rev": "CCATAGAAGCCTTGGTTGAACATG"},
    "NB61": {"fwd": "AGAGGGTACTATGTGCCTCAGCAC", "rev": "GTGCTGAGGCACATAGTACCCTCT"},
    "NB62": {"fwd": "CACCCACACTTACTTCAGGACGTA", "rev": "TACGTCCTGAAGTAAGTGTGGGTG"},
    "NB63": {"fwd": "TTCTGAAGTTCCTGGGTCTTGAAC", "rev": "GTTCAAGACCCAGGAACTTCAGAA"},
    "NB64": {"fwd": "GACAGACACCGTTCATCGACTTTC", "rev": "GAAAGTCGATGAACGGTGTCTGTC"},
    "NB65": {"fwd": "TTCTCAGTCTTCCTCCAGACAAGG", "rev": "CCTTGTCTGGAGGAAGACTGAGAA"},
    "NB66": {"fwd": "CCGATCCTTGTGGCTTCTAACTTC", "rev": "GAAGTTAGAAGCCACAAGGATCGG"},
    "NB67": {"fwd": "GTTTGTCATACTCGTGTGCTCACC", "rev": "GGTGAGCACACGAGTATGACAAAC"},
    "NB68": {"fwd": "GAATCTAAGCAAACACGAAGGTGG", "rev": "CCACCTTCGTGTTTGCTTAGATTC"},
    "NB69": {"fwd": "TACAGTCCGAGCCTCATGTGATCT", "rev": "AGATCACATGAGGCTCGGACTGTA"},
    "NB70": {"fwd": "ACCGAGATCCTACGAATGGAGTGT", "rev": "ACACTCCATTCGTAGGATCTCGGT"},
    "NB71": {"fwd": "CCTGGGAGCATCAGGTAGTAACAG", "rev": "CTGTTACTACCTGATGCTCCCAGG"},
    "NB72": {"fwd": "TAGCTGACTGTCTTCCATACCGAC", "rev": "GTCGGTATGGAAGACAGTCAGCTA"},
    "NB73": {"fwd": "AAGAAACAGGATGACAGAACCCTC", "rev": "GAGGGTTCTGTCATCCTGTTTCTT"},
    "NB74": {"fwd": "TACAAGCATCCCAACACTTCCACT", "rev": "AGTGGAAGTGTTGGGATGCTTGTA"},
    "NB75": {"fwd": "GACCATTGTGATGAACCCTGTTGT", "rev": "ACAACAGGGTTCATCACAATGGTC"},
    "NB76": {"fwd": "ATGCTTGTTACATCAACCCTGGAC", "rev": "GTCCAGGGTTGATGTAACAAGCAT"},
    "NB77": {"fwd": "CGACCTGTTTCTCAGGGATACAAC", "rev": "GTTGTATCCCTGAGAAACAGGTCG"},
    "NB78": {"fwd": "AACAACCGAACCTTTGAATCAGAA", "rev": "TTCTGATTCAAAGGTTCGGTTGTT"},
    "NB79": {"fwd": "TCTCGGAGATAGTTCTCACTGCTG", "rev": "CAGCAGTGAGAACTATCTCCGAGA"},
    "NB80": {"fwd": "CGGATGAACATAGGATAGCGATTC", "rev": "GAATCGCTATCCTATGTTCATCCG"},
    "NB81": {"fwd": "CCTCATCTTGTGAAGTTGTTTCGG", "rev": "CCGAAACAACTTCACAAGATGAGG"},
    "NB82": {"fwd": "ACGGTATGTCGAGTTCCAGGACTA", "rev": "TAGTCCTGGAACTCGACATACCGT"},
    "NB83": {"fwd": "TGGCTTGATCTAGGTAAGGTCGAA", "rev": "TTCGACCTTACCTAGATCAAGCCA"},
    "NB84": {"fwd": "GTAGTGGACCTAGAACCTGTGCCA", "rev": "TGGCACAGGTTCTAGGTCCACTAC"},
    "NB85": {"fwd": "AACGGAGGAGTTAGTTGGATGATC", "rev": "GATCATCCAACTAACTCCTCCGTT"},
    "NB86": {"fwd": "AGGTGATCCCAACAAGCGTAAGTA", "rev": "TACTTACGCTTGTTGGGATCACCT"},
    "NB87": {"fwd": "TACATGCTCCTGTTGTTAGGGAGG", "rev": "CCTCCCTAACAACAGGAGCATGTA"},
    "NB88": {"fwd": "TCTTCTACTACCGATCCGAAGCAG", "rev": "CTGCTTCGGATCGGTAGTAGAAGA"},
    "NB89": {"fwd": "ACAGCATCAATGTTTGGCTAGTTG", "rev": "CAACTAGCCAAACATTGATGCTGT"},
    "NB90": {"fwd": "GATGTAGAGGGTACGGTTTGAGGC", "rev": "GCCTCAAACCGTACCCTCTACATC"},
    "NB91": {"fwd": "GGCTCCATAGGAACTCACGCTACT", "rev": "AGTAGCGTGAGTTCCTATGGAGCC"},
    "NB92": {"fwd": "TTGTGAGTGGAAAGATACAGGACC", "rev": "GGTCCTGTATCTTTCCACTCACAA"},
    "NB93": {"fwd": "AGTTTCCATCACTTCAGACTTGGG", "rev": "CCCAAGTCTGAAGTGATGGAAACT"},
    "NB94": {"fwd": "GATTGTCCTCAAACTGCCACCTAC", "rev": "GTAGGTGGCAGTTTGAGGACAATC"},
    "NB95": {"fwd": "CCTGTCTGGAAGAAGAATGGACTT", "rev": "AAGTCCATTCTTCTTCCAGACAGG"},
    "NB96": {"fwd": "CTGAACGGTCATAGAGTCCACCAT", "rev": "ATGGTGGACTCTATGACCGTTCAG"},
}

# Sorted list of all NB labels in numeric order (NB01 … NB96)
_ALL_NB_LABELS: List[str] = sorted(ALL_BARCODES.keys(), key=lambda x: int(x[2:]))

# ---------------------------------------------------------------------------
# Sequence utilities
# ---------------------------------------------------------------------------

def _revcomp(seq: str) -> str:
    """Return the reverse complement of a DNA sequence."""
    return seq.upper().translate(str.maketrans("ACGT", "TGCA"))[::-1]


# ---------------------------------------------------------------------------
# History loading
# ---------------------------------------------------------------------------

def load_history(excel_path: str) -> Tuple[Dict[str, str], Set[Tuple[str, str]]]:
    """
    Load barcode assignment history from an Excel file.

    Expected columns: "4-bp Overhang", "Barcode label"
    Rows that also contain a "Part/Fragment" or similar grouping column allow
    reconstruction of used (oh5_nb, oh3_nb) pairs.  We detect this by looking
    for a column whose values look like part-fragment identifiers.

    Returns
    -------
    oh_to_nb   : dict mapping overhang sequence → NB label (e.g. "GGAG" → "NB25")
    used_pairs : set of (nb_oh5, nb_oh3) tuples already committed in history
    """
    df = pd.read_excel(excel_path)

    # Normalise column names (strip whitespace, case-insensitive match)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}

    oh_col = col_map.get("4-bp overhang") or col_map.get("overhang") or col_map.get("4bp overhang")
    nb_col = col_map.get("barcode label") or col_map.get("barcode") or col_map.get("nb label")

    if oh_col is None or nb_col is None:
        raise ValueError(
            f"History Excel must have columns '4-bp Overhang' and 'Barcode label'. "
            f"Found: {list(df.columns)}"
        )

    # Build oh_to_nb from rows where both cells are filled
    oh_to_nb: Dict[str, str] = {}
    for _, row in df.iterrows():
        oh  = str(row[oh_col]).strip().upper()
        nb  = str(row[nb_col]).strip()
        # pandas represents missing values as the string "nan" (lowercase)
        if oh and nb and oh.upper() != "NAN" and nb.upper() != "NAN" and len(oh) == 4:
            oh_to_nb[oh] = nb

    # Reconstruct used pairs – look for an F/R direction column
    used_pairs: Set[Tuple[str, str]] = set()
    dir_col = col_map.get("direction") or col_map.get("end") or col_map.get("f/r")
    frag_col = (col_map.get("fragment") or col_map.get("part/fragment")
                or col_map.get("part") or col_map.get("frag"))

    if dir_col and frag_col:
        # Group by fragment; within each group find F (oh5) and R (oh3) barcodes
        for frag_id, grp in df.groupby(df[frag_col]):
            f_rows = grp[grp[dir_col].astype(str).str.upper().str.startswith("F")]
            r_rows = grp[grp[dir_col].astype(str).str.upper().str.startswith("R")]
            for _, fr in f_rows.iterrows():
                for _, rr in r_rows.iterrows():
                    nb5 = str(fr[nb_col]).strip()
                    nb3 = str(rr[nb_col]).strip()
                    if nb5 != "NAN" and nb3 != "NAN":
                        used_pairs.add((nb5, nb3))

    return oh_to_nb, used_pairs


# ---------------------------------------------------------------------------
# Barcode assignment
# ---------------------------------------------------------------------------

def assign_barcodes(
    parts: List[Part],
    oh_to_nb: Dict[str, str],
    used_pairs: Set[Tuple[str, str]],
) -> Dict[Tuple[str, int], Dict[str, str]]:
    """
    Assign NB barcodes to each fragment's oh5 and oh3 overhangs.

    Parameters
    ----------
    parts       : list of Part objects from split_gene()
    oh_to_nb    : existing overhang→NB mapping from history (modified in-place)
    used_pairs  : existing (nb_oh5, nb_oh3) pairs already in use (modified in-place)

    Returns
    -------
    assignments : dict mapping (part_label, frag_index) →
                  {"oh5": str, "oh3": str, "nb_oh5": str, "nb_oh3": str,
                   "oh5_reused": bool, "oh3_reused": bool}
    """
    # Count usage frequency of each NB across history
    usage: Counter = Counter()
    for nb in oh_to_nb.values():
        usage[nb] += 1

    def _least_used_nb(exclude: Optional[str] = None) -> str:
        """Return the NB label with the fewest assignments so far."""
        best_nb = None
        best_count = float("inf")
        for nb in _ALL_NB_LABELS:
            if nb == exclude:
                continue
            c = usage.get(nb, 0)
            if c < best_count:
                best_count = c
                best_nb = nb
        return best_nb  # type: ignore[return-value]

    assignments: Dict[Tuple[str, int], Dict[str, str]] = {}

    for part in parts:
        n_frags = len(part.fragments)
        for frag_idx, frag in enumerate(part.fragments):
            oh5 = frag.overhang_5
            oh3 = frag.overhang_3

            # ---- Assign oh5 barcode ----
            oh5_reused = oh5 in oh_to_nb
            if oh5_reused:
                nb_oh5 = oh_to_nb[oh5]
            else:
                nb_oh5 = _least_used_nb()
                oh_to_nb[oh5] = nb_oh5
                usage[nb_oh5] = usage.get(nb_oh5, 0) + 1

            # ---- Assign oh3 barcode ----
            oh3_reused = oh3 in oh_to_nb
            if oh3_reused:
                nb_oh3 = oh_to_nb[oh3]
            else:
                nb_oh3 = _least_used_nb(exclude=nb_oh5)
                oh_to_nb[oh3] = nb_oh3
                usage[nb_oh3] = usage.get(nb_oh3, 0) + 1

            # ---- Ensure the (oh5, oh3) pair is unique ----
            pair = (nb_oh5, nb_oh3)
            if pair in used_pairs:
                # Try replacing oh3 barcode with the next-least-used NB
                # that produces a unique pair
                for nb_candidate in sorted(_ALL_NB_LABELS, key=lambda x: usage.get(x, 0)):
                    if nb_candidate == nb_oh5:
                        continue
                    if (nb_oh5, nb_candidate) not in used_pairs:
                        # Undo the previous oh3 assignment if it was new
                        if not oh3_reused:
                            usage[nb_oh3] = max(0, usage.get(nb_oh3, 1) - 1)
                        nb_oh3 = nb_candidate
                        oh_to_nb[oh3] = nb_oh3
                        usage[nb_oh3] = usage.get(nb_oh3, 0) + 1
                        oh3_reused = False  # it's a new assignment now
                        pair = (nb_oh5, nb_oh3)
                        break

            used_pairs.add(pair)

            assignments[(part.label, frag.index)] = {
                "oh5": oh5,
                "oh3": oh3,
                "nb_oh5": nb_oh5,
                "nb_oh3": nb_oh3,
                "oh5_reused": oh5_reused,
                "oh3_reused": oh3_reused,
            }

    return assignments


# ---------------------------------------------------------------------------
# .dna file builders
# ---------------------------------------------------------------------------

# Constant flanking sequences for the barcode adapter strands
_TOP_UPSTREAM   = "AAGGTTAA"    # 8 nt
_TOP_DOWNSTREAM = "CAGCACCT"    # 8 nt
_BOT_UPSTREAM   = "AGGTGCTG"    # 8 nt  (revcomp of _TOP_DOWNSTREAM)
_BOT_DOWNSTREAM = "TTAACCTT"    # 8 nt  (revcomp of _TOP_UPSTREAM)
_BOT_TAIL       = "AGCAAT"      # 6 nt  (the recessed 3' tail / sticky)
_ATTGCT         = "ATTGCT"      # 6 nt  prefix for BottomStrand & adapter files


def build_top_strand_bytes(nb_label: str, is_circular: bool = False) -> bytes:
    """
    Build the NBXX_TopStrand.dna file bytes.

    Sequence (40 nt):
        AAGGTTAA + barcode_fwd_24nt + CAGCACCT

    Features (1-based):
        upstream_dsDNA   : 1–8,  #ff99cc
        nanopore_barcode : 9–32, #800080
        downstream_dsDNA : 33–40, #ff99cc
    """
    bc = ALL_BARCODES[nb_label]
    seq = _TOP_UPSTREAM + bc["fwd"] + _TOP_DOWNSTREAM
    assert len(seq) == 40, f"TopStrand length error: {len(seq)}"

    features = [
        make_feature("upstream_dsDNA",            start=0,  end=8,  strand=1,
                     ftype="misc_feature", color="#ff99cc"),
        make_feature(f"nanopore_barcode_{nb_label}_fwd", start=8,  end=32, strand=1,
                     ftype="misc_feature", color="#800080"),
        make_feature("downstream_dsDNA",          start=32, end=40, strand=1,
                     ftype="misc_feature", color="#ff99cc"),
    ]

    return build_dna_bytes(
        sequence=seq,
        is_circular=is_circular,
        features=features,
        name=f"{nb_label}_TopStrand",
        add_phospho_ends=False,
        upstream_sticky=0,
        downstream_sticky=0,
        upstream_mod="Unmodified",
        downstream_mod="Unmodified",
    )


def build_bottom_strand_bytes(
    nb_label: str,
    overhang_4bp: str,
    is_circular: bool = False,
) -> bytes:
    """
    Build the NBXX_BottomStrand.dna file bytes.

    Sequence (50 nt):
        revcomp(oh_4bp) + AGGTGCTG + barcode_rev_24nt + TTAACCTT + AGCAAT

    Feature positions (1-based in spec; stored 0-based here):
        upstream_dsDNA         : pos  6–12  → 0-based  5–11  (end=12)
        nanopore_barcode_rev   : pos 13–36  → 0-based 12–35  (end=36)
        downstream_dsDNA       : pos 37–50  → 0-based 36–50  (end=50)
    Note: the first 5 nt are the sticky revcomp(oh) + 1 nt of the upstream flank.
    Per spec: upstream_dsDNA range 6-12 means nt 6..12 (1-based), i.e. the last
    2 nt of revcomp(oh) + first 6 nt of _BOT_UPSTREAM (AGGTGCTG without last 2).
    We follow the spec literally: range 6–12 = 0-based [5,12).
    """
    bc = ALL_BARCODES[nb_label]
    oh_rc = _revcomp(overhang_4bp)  # 4 nt
    seq = oh_rc + _BOT_UPSTREAM + bc["rev"] + _BOT_DOWNSTREAM + _BOT_TAIL
    assert len(seq) == 50, f"BottomStrand length error: {len(seq)}"

    features = [
        make_feature("upstream_dsDNA",            start=5,  end=12, strand=1,
                     ftype="misc_feature", color="#ff99cc"),
        make_feature(f"nanopore_barcode_{nb_label}_rev", start=12, end=36, strand=1,
                     ftype="misc_feature", color="#800080"),
        make_feature("downstream_dsDNA",          start=36, end=50, strand=1,
                     ftype="misc_feature", color="#ff99cc"),
    ]

    return build_dna_bytes(
        sequence=seq,
        is_circular=is_circular,
        features=features,
        name=f"{nb_label}_BottomStrand",
        add_phospho_ends=False,
        upstream_sticky=0,
        downstream_sticky=0,
        upstream_mod="Unmodified",
        downstream_mod="Unmodified",
    )


def build_adapter_dna_bytes(nb_label: str, overhang_4bp: str) -> bytes:
    """
    Build the XXXX_NBXX.dna annealed adapter file bytes.
    XXXX = revcomp(overhang_4bp) — the label reflects the sticky end exposed.

    Top strand (50 nt):
        ATTGCT + AAGGTTAA + barcode_fwd_24nt + CAGCACCT + oh_4bp

    Stickiness: upstream=-6 (ATTGCT is the recessed 5' tail contributing -6),
                downstream=+4 (oh_4bp is the 3' overhang contributing +4).
    End mods: FivePrimePhosphorylated on both ends.

    Features (multi-segment spec):
    1. upstream_nanopore_barcode_sequence (misc_feature) — two segments:
         "3' overhang"  : pos 1–6   (0-based 0–5,   i.e. ATTGCT)        color #ff99cc
         "dsDNA"        : pos 7–14  (0-based 6–13,  i.e. AAGGTTAA)      color #993366
       Wait — spec says range 1–7 and 8–14.  Adjusting: ATTGCT is 6 nt (pos 1-6),
       then AAGGTTAA is 8 nt (pos 7-14). Spec says 1–7 and 8–14, so segment 1 = 7 nt.
       That implies ATTGCT (6) + first 1 of AAGGTTAA = 7 nt for "3' overhang".
       Segment 2 "dsDNA" = pos 8–14 = remaining 7 of AAGGTTAA (7 nt).
       We follow the spec: segment 1 range 1-7 (0-based 0-6), segment 2 range 8-14 (0-based 7-13).

    2. NBXX (misc_feature): range 15–38 (0-based 14–37) = barcode_fwd_24nt, color #800080

    3. downstream_nanopore_barcode_sequence (misc_feature) — two segments:
         "dsDNA"         : range 39–45 (0-based 38–44) = CAGCACCT[0:7]  color #993366
         "3' dT overhang": range 46–46 (0-based 45–45) = last nt of CAGCACCT color #ff99cc
       Followed by oh_4bp which is the sticky end (pos 47-50, 0-based 46-49) — not a feature.
       Wait — CAGCACCT is 8 nt (pos 39-46), then oh_4bp is pos 47-50.
       Spec says downstream has "dsDNA" 39-45 and "3' dT overhang" 46-46.
       So pos 47-50 (oh_4bp) is the sticky 3' overhang and is not annotated as a feature.
    """
    bc = ALL_BARCODES[nb_label]
    oh = overhang_4bp.upper()
    seq = _ATTGCT + _TOP_UPSTREAM + bc["fwd"] + _TOP_DOWNSTREAM + oh
    assert len(seq) == 50, f"Adapter length error: {len(seq)}"

    # upstream multi-seg: ATTGCT(0-5) + A(6) = 0-6 "3' overhang"; AGGTTAA(7-13) = "dsDNA"
    # downstream multi-seg: CAGCACC(38-44) = "dsDNA"; T(45) = "3' dT overhang"
    features = [
        make_multiseg_feature(
            name="upstream_nanopore_barcode_sequence",
            ftype="misc_feature",
            segments=[
                {"start": 0,  "end": 6,  "color": "#ff99cc", "name": "3' overhang"},
                {"start": 7,  "end": 13, "color": "#993366", "name": "dsDNA"},
            ],
        ),
        make_feature(
            name=nb_label,
            start=14, end=38,
            strand=1,
            ftype="misc_feature",
            color="#800080",
        ),
        make_multiseg_feature(
            name="downstream_nanopore_barcode_sequence",
            ftype="misc_feature",
            segments=[
                {"start": 38, "end": 44, "color": "#993366", "name": "dsDNA"},
                {"start": 45, "end": 45, "color": "#ff99cc", "name": "3' dT overhang"},
            ],
        ),
    ]

    return build_dna_bytes(
        sequence=seq,
        is_circular=False,
        features=features,
        name=f"{_revcomp(oh)}_{nb_label}",
        add_phospho_ends=False,
        upstream_sticky=-6,
        downstream_sticky=4,
        upstream_mod="FivePrimePhosphorylated",
        downstream_mod="FivePrimePhosphorylated",
    )


def build_ligation_bytes(
    frag: Fragment,
    nb_oh5_label: str,
    nb_oh3_label: str,
    is_first_of_part_a: bool = False,
    is_last_of_part_e: bool = False,
) -> bytes:
    """
    Build the Ligation.dna file bytes representing the full nanopore sequencing
    ligation product for one fragment.

    Top strand layout (5'→3'):
        [36 nt]  TTTTTTTT + CCTGTACTTCGTTCAGTTACGT + ATTGCT   ← left native adapter
        [40 nt]  AAGGTTAA + NB_oh5_fwd_24nt + CAGCACCT        ← TopStrand of oh5 barcode
        [4 nt]   oh5_4nt                                        ← oh5 overhang junction
        [5 nt]   ATATC (only if is_first_of_part_a)            ← EcoRV site
        [N nt]   frag.sequence                                  ← gene insert
        [5 nt]   ATATC (only if is_last_of_part_e)             ← EcoRV site (rev end)
        [50 nt]  revcomp(oh3_4nt)+AGGTGCTG+NB_oh3_rev_24nt+TTAACCTT+AGCAAT ← BottomStrand
        [30 nt]  ACGTAACTGAACGAAGTACAGG + AAAAAAAA            ← right native adapter

    Stickiness: upstream=+8 (poly-dT), downstream=-8 (poly-A on right end).
    End modifications: Unmodified (the native adapters are pre-ligated).
    """
    bc_oh5 = ALL_BARCODES[nb_oh5_label]
    bc_oh3 = ALL_BARCODES[nb_oh3_label]
    oh5 = frag.overhang_5
    oh3 = frag.overhang_3

    # ---- Build each section ----
    left_native   = "TTTTTTTT" + "CCTGTACTTCGTTCAGTTACGT" + "ATTGCT"   # 36 nt
    top_strand_40 = _TOP_UPSTREAM + bc_oh5["fwd"] + _TOP_DOWNSTREAM     # 40 nt
    oh5_junction  = oh5.upper()                                           # 4 nt
    ecorv_fwd     = "ATATC" if is_first_of_part_a else ""                # 5 or 0 nt
    gene_insert   = frag.sequence.upper()
    ecorv_rev     = "ATATC" if is_last_of_part_e else ""                 # 5 or 0 nt
    bot_strand_50 = (_revcomp(oh3) + _BOT_UPSTREAM
                     + bc_oh3["rev"] + _BOT_DOWNSTREAM + _BOT_TAIL)     # 50 nt
    right_native  = "ACGTAACTGAACGAAGTACAGG" + "AAAAAAAA"               # 30 nt

    seq = (left_native + top_strand_40 + oh5_junction
           + ecorv_fwd + gene_insert + ecorv_rev
           + bot_strand_50 + right_native)

    # ---- Feature positions ----
    p = 0
    left_native_start   = p;               p += len(left_native)    # 0–35
    top_strand_start    = p;               p += len(top_strand_40)  # 36–75
    oh5_junc_start      = p;              p += len(oh5_junction)   # 76–79
    ecorv_fwd_start     = p if is_first_of_part_a else None
    if is_first_of_part_a:
        p += 5
    gene_insert_start   = p;               p += len(gene_insert)
    if is_last_of_part_e:
        ecorv_rev_start = p
        p += 5
    else:
        ecorv_rev_start = None
    bot_strand_start    = p;               p += len(bot_strand_50) # 50 nt
    right_native_start  = p;              p += len(right_native)  # 30 nt

    gene_insert_end     = gene_insert_start + len(gene_insert)
    bot_strand_end      = bot_strand_start + len(bot_strand_50)
    right_native_end    = right_native_start + len(right_native)

    features = [
        make_feature(
            "Nanopore_Native_Adapter_L",
            start=left_native_start, end=left_native_start + len(left_native),
            strand=1, ftype="misc_feature", color="#008000",
        ),
        make_feature(
            f"NB_oh5_barcode ({nb_oh5_label})",
            start=top_strand_start, end=top_strand_start + len(top_strand_40),
            strand=1, ftype="misc_feature", color="#800080",
        ),
        make_feature(
            f"oh5_junction ({oh5})",
            start=oh5_junc_start, end=oh5_junc_start + 4,
            strand=1, ftype="misc_feature", color="#f5a623",
        ),
    ]

    if ecorv_fwd_start is not None:
        features.append(make_feature(
            "EcoRV",
            start=ecorv_fwd_start, end=ecorv_fwd_start + 5,
            strand=1, ftype="misc_feature", color="#9b59b6",
        ))

    features.append(make_feature(
        f"{frag.part_label}.{frag.index} gene insert",
        start=gene_insert_start, end=gene_insert_end,
        strand=1, ftype="CDS", color="#0000ff",
    ))

    if ecorv_rev_start is not None:
        features.append(make_feature(
            "EcoRV",
            start=ecorv_rev_start, end=ecorv_rev_start + 5,
            strand=1, ftype="misc_feature", color="#9b59b6",
        ))

    features.append(make_feature(
        f"NB_oh3_barcode ({nb_oh3_label})",
        start=bot_strand_start, end=bot_strand_end,
        strand=1, ftype="misc_feature", color="#800080",
    ))
    features.append(make_feature(
        "Nanopore_Native_Adapter_R",
        start=right_native_start, end=right_native_end,
        strand=1, ftype="misc_feature", color="#008000",
    ))

    return build_dna_bytes(
        sequence=seq,
        is_circular=False,
        features=features,
        name=f"{frag.part_label}.{frag.index} Ligation",
        add_phospho_ends=False,
        upstream_sticky=8,
        downstream_sticky=-8,
        upstream_mod="Unmodified",
        downstream_mod="Unmodified",
    )


# ---------------------------------------------------------------------------
# Master generation function
# ---------------------------------------------------------------------------

def generate_all_barcode_files(
    parts: List[Part],
    oh_to_nb: Dict[str, str],
    used_pairs: Set[Tuple[str, str]],
) -> Dict[str, bytes]:
    """
    Generate all barcode .dna files for every fragment across all Parts.

    Performs barcode assignment, then for each fragment produces:
      - Part X/X.N/NBXX_TopStrand.dna          (for oh5 barcode)
      - Part X/X.N/NBXX_BottomStrand.dna        (for oh5 barcode, using oh5 seq)
      - Part X/X.N/XXXX_NBXX.dna               (annealed adapter for oh5)
      - Part X/X.N/NBXX_TopStrand.dna          (for oh3 barcode, if different)
      - Part X/X.N/NBXX_BottomStrand.dna        (for oh3 barcode, using oh3 seq)
      - Part X/X.N/XXXX_NBXX.dna               (annealed adapter for oh3)
      - Part X/X.N/Ligation.dna

    Files for a given NB label are deduplicated: the TopStrand is identical
    regardless of which fragment references it, so we only write it once per
    unique NB label used in a given fragment folder.

    Returns
    -------
    dict mapping relative path string → file bytes
    """
    assignments = assign_barcodes(parts, oh_to_nb, used_pairs)
    files: Dict[str, bytes] = {}

    n_frags_per_part = {part.label: len(part.fragments) for part in parts}

    for part in parts:
        n_frags = len(part.fragments)
        for frag_idx, frag in enumerate(part.fragments):
            is_a1     = (part.label == "A" and frag_idx == 0)
            is_e_last = (part.label == "E" and frag_idx == n_frags - 1)

            key = (part.label, frag.index)
            asgn = assignments[key]
            nb_oh5 = asgn["nb_oh5"]
            nb_oh3 = asgn["nb_oh3"]
            oh5    = asgn["oh5"]
            oh3    = asgn["oh3"]

            folder = f"Part_{part.label}/{part.label}.{frag.index}"

            # ---- oh5 barcode files ----
            top5_path = f"{folder}/{nb_oh5}_TopStrand.dna"
            if top5_path not in files:
                files[top5_path] = build_top_strand_bytes(nb_oh5)

            bot5_path = f"{folder}/{nb_oh5}_BottomStrand.dna"
            if bot5_path not in files:
                files[bot5_path] = build_bottom_strand_bytes(nb_oh5, oh5)

            # XXXX = revcomp(oh5) — the label on the adapter file
            oh5_rc = _revcomp(oh5)
            adp5_path = f"{folder}/{oh5_rc}_{nb_oh5}.dna"
            if adp5_path not in files:
                files[adp5_path] = build_adapter_dna_bytes(nb_oh5, oh5)

            # ---- oh3 barcode files ----
            top3_path = f"{folder}/{nb_oh3}_TopStrand.dna"
            if top3_path not in files:
                files[top3_path] = build_top_strand_bytes(nb_oh3)

            bot3_path = f"{folder}/{nb_oh3}_BottomStrand.dna"
            if bot3_path not in files:
                files[bot3_path] = build_bottom_strand_bytes(nb_oh3, oh3)

            oh3_rc = _revcomp(oh3)
            adp3_path = f"{folder}/{oh3_rc}_{nb_oh3}.dna"
            if adp3_path not in files:
                files[adp3_path] = build_adapter_dna_bytes(nb_oh3, oh3)

            # ---- Ligation product ----
            lig_path = f"{folder}/Ligation.dna"
            files[lig_path] = build_ligation_bytes(frag, nb_oh5, nb_oh3, is_a1, is_e_last)

    return files


# ---------------------------------------------------------------------------
# Excel summary
# ---------------------------------------------------------------------------

def build_barcode_excel(
    parts: List[Part],
    assignments: Dict[Tuple[str, int], Dict[str, str]],
) -> bytes:
    """
    Build a barcode-assignment summary Excel file in memory and return bytes.

    Columns: Fragment, Part, oh5, NB_oh5, oh5_status, oh3, NB_oh3, oh3_status
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Barcode Assignments"

    headers = ["Fragment", "Part", "oh5", "NB_oh5", "oh5 (new/reused)",
               "oh3", "NB_oh3", "oh3 (new/reused)"]
    ws.append(headers)

    # Header styling
    header_font  = Font(name="Calibri", bold=True, size=11)
    header_fill  = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_font = Font(name="Calibri", size=11)

    for part in parts:
        for frag in part.fragments:
            key  = (part.label, frag.index)
            asgn = assignments.get(key)
            if asgn is None:
                continue
            ws.append([
                f"{part.label}.{frag.index}",
                part.label,
                asgn["oh5"],
                asgn["nb_oh5"],
                "reused" if asgn["oh5_reused"] else "new",
                asgn["oh3"],
                asgn["nb_oh3"],
                "reused" if asgn["oh3_reused"] else "new",
            ])
            for cell in ws[ws.max_row]:
                cell.font = row_font

    # Auto-fit column widths (approximate)
    col_widths = [12, 8, 8, 10, 16, 8, 10, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# ZIP helpers
# ---------------------------------------------------------------------------

def build_barcode_zip(files: Dict[str, bytes]) -> bytes:
    """Return a ZIP archive (bytes) of the provided {path: bytes} dict."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path, data in sorted(files.items()):
            zf.writestr(path, data)
    return buf.getvalue()


def build_combined_zip(
    gg_output_dir: str,
    barcode_files: Dict[str, bytes],
) -> bytes:
    """
    Combine the Golden Gate output directory files with the in-memory barcode
    files into a single ZIP archive.

    Parameters
    ----------
    gg_output_dir  : path to the directory written by run_pipeline()
    barcode_files  : dict from generate_all_barcode_files()
    """
    from pathlib import Path as _Path
    gg_dir = _Path(gg_output_dir)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Add GG files
        for file in sorted(gg_dir.rglob("*")):
            if file.is_file():
                zf.write(file, file.relative_to(gg_dir))
        # Add barcode files (placed inside a "Barcodes" sub-folder)
        for rel_path, data in sorted(barcode_files.items()):
            zf.writestr(f"Barcodes/{rel_path}", data)
    return buf.getvalue()
